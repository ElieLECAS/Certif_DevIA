def cu_performance_export_excel(request):
    """Vue pour exporter les données de performance des CU en Excel avec graphique"""
    from django.http import HttpResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from .services import get_cu_performance_from_db, get_available_dates_from_db, get_available_cu_types_from_db, format_session_for_display
    from datetime import datetime, timedelta
    import io
    import zipfile
    import base64
    
    # Récupérer les paramètres de la requête
    cu_type = request.GET.get('cu_type', '')
    date_debut = request.GET.get('date_debut', '')
    date_fin = request.GET.get('date_fin', '')
    
    # Validation des paramètres
    if not cu_type:
        return HttpResponse("Type de CU requis", status=400)
    
    if not date_debut or not date_fin:
        return HttpResponse("Dates de début et fin requises", status=400)
    
    try:
        date_debut = datetime.strptime(date_debut, '%Y-%m-%d').date()
        date_fin = datetime.strptime(date_fin, '%Y-%m-%d').date()
    except ValueError:
        return HttpResponse("Format de date invalide", status=400)
    
    if date_debut > date_fin:
        return HttpResponse("La date de début doit être antérieure à la date de fin", status=400)
    
    # Créer le classeur Excel
    wb = Workbook()
    ws = wb.active
    ws.title = f"Performance CU {cu_type}"
    
    # Configuration des styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    # En-têtes des colonnes
    headers = [
        'Date', 'CU ID', 'Total Pièces', 'Durée Production (h)', 
        'Machine Start', 'Première Pièce', 'Dernière Pièce', 'Machine Stop',
        'Temps Attente (h)', 'Temps Arrêt Volontaire (h)', 
        'Temps Production Effectif (h)', 'Taux Occupation (%)', 
        'Taux Attente (%)', 'Taux Arrêt Volontaire (%)'
    ]
    
    # Écrire les en-têtes
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = border
    
    # Récupérer les données pour la plage de dates et stocker les sessions pour les graphiques
    current_date = date_debut
    row = 2
    sessions_with_data = []  # Pour stocker les sessions qui ont des données
    
    while current_date <= date_fin:
        try:
            # Récupérer les sessions pour cette date
            sessions = get_cu_performance_from_db(
                cu_type=cu_type,
                date_production=current_date
            )
            
            if sessions.exists():
                for session in sessions:
                    # Formater les données pour l'affichage
                    results = format_session_for_display(session)
                    
                    # Stocker la session pour générer les graphiques
                    sessions_with_data.append({
                        'date': current_date,
                        'results': results
                    })
                    
                    # Écrire les données dans Excel
                    data_row = [
                        current_date.strftime('%d/%m/%Y'),
                        results.get('CU_ID', ''),
                        results.get('TotalPieces', 0),
                        round(results.get('DureeProduction', 0), 2),
                        results.get('PremierMachineStart', '').strftime('%H:%M') if results.get('PremierMachineStart') else 'N/A',
                        results.get('PremierePiece', '').strftime('%H:%M') if results.get('PremierePiece') else 'N/A',
                        results.get('DernierePiece', '').strftime('%H:%M') if results.get('DernierePiece') else 'N/A',
                        results.get('DernierMachineStop', '').strftime('%H:%M') if results.get('DernierMachineStop') else 'N/A',
                        round(results.get('TempsAttente', 0), 2),
                        round(results.get('TempsArretVolontaire', 0), 2),
                        round(results.get('TempsProductionEffectif', 0), 2),
                        round(results.get('TauxOccupation', 0), 1),
                        round(results.get('TauxAttente', 0), 1),
                        round(results.get('TauxArretVolontaire', 0), 1)
                    ]
                    
                    for col, value in enumerate(data_row, 1):
                        cell = ws.cell(row=row, column=col, value=value)
                        cell.border = border
                        cell.alignment = center_alignment
                    
                    row += 1
            else:
                # Ligne vide pour les jours sans données
                data_row = [
                    current_date.strftime('%d/%m/%Y'),
                    'Aucune donnée',
                    '', '', '', '', '', '', '', '', '', '', '', ''
                ]
                
                for col, value in enumerate(data_row, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.border = border
                    cell.alignment = center_alignment
                    if col == 2:  # Colonne "Aucune donnée"
                        cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
                
                row += 1
                
        except Exception as e:
            print(f"Erreur pour la date {current_date}: {str(e)}")
        
        current_date += timedelta(days=1)
    
    # Ajuster la largeur des colonnes
    for col in range(1, len(headers) + 1):
        column_letter = get_column_letter(col)
        ws.column_dimensions[column_letter].width = 15
    
    # Ajouter des statistiques en bas
    if row > 2:  # S'il y a des données
        row += 2  # Laisser une ligne vide
        
        # En-tête des statistiques
        stats_cell = ws.cell(row=row, column=1, value="STATISTIQUES PÉRIODE")
        stats_cell.font = Font(bold=True, size=12)
        stats_cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        ws.merge_cells(f'A{row}:N{row}')
        row += 1
    
    # Créer un fichier ZIP contenant l'Excel et les graphiques
    zip_buffer = io.BytesIO()
    
    try:
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            # Ajouter le fichier Excel au ZIP
            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            
            excel_filename = f"Performance_CU_{cu_type}_{date_debut.strftime('%Y%m%d')}_{date_fin.strftime('%Y%m%d')}.xlsx"
            zip_file.writestr(excel_filename, excel_buffer.getvalue())
            
            # Générer et ajouter les graphiques de chronologie pour chaque session
            chart_count = 0
            for session_data in sessions_with_data:
                try:
                    # Générer le graphique de chronologie pour cette session
                    fallback_charts = generate_fallback_charts(session_data['results'])
                    
                    if fallback_charts and fallback_charts.get('timeline_chart'):
                        # Extraire l'image base64 et la convertir en bytes
                        timeline_data = fallback_charts['timeline_chart']
                        
                        # Enlever le préfixe "data:image/png;base64,"
                        if timeline_data.startswith('data:image/png;base64,'):
                            timeline_data = timeline_data[len('data:image/png;base64,'):]
                        
                        # Décoder l'image
                        image_bytes = base64.b64decode(timeline_data)
                        
                        # Nom du fichier image
                        chart_filename = f"Chronologie_{cu_type}_{session_data['date'].strftime('%Y%m%d')}_{session_data['results'].get('CU_ID', 'CU')}.png"
                        
                        # Ajouter l'image au ZIP
                        zip_file.writestr(chart_filename, image_bytes)
                        chart_count += 1
                        
                except Exception as e:
                    print(f"Erreur lors de la génération du graphique pour {session_data['date']}: {str(e)}")
                    continue
            
            # Ajouter un fichier README avec les informations sur l'export
            readme_content = f"""Export des données de performance CU {cu_type}
Période: du {date_debut.strftime('%d/%m/%Y')} au {date_fin.strftime('%d/%m/%Y')}
Date d'export: {datetime.now().strftime('%d/%m/%Y à %H:%M')}

Contenu du fichier ZIP:
- {excel_filename}: Données de performance au format Excel
- {chart_count} graphique(s) de chronologie au format PNG

Les graphiques de chronologie montrent:
- Périodes de production (vert)
- Périodes d'attente (jaune)
- Périodes d'arrêt (rouge)
- Points bleus: pièces produites

Généré par le Dashboard Suivi Proferm
"""
            zip_file.writestr("README.txt", readme_content.encode('utf-8'))
        
        # Préparer la réponse HTTP
        zip_buffer.seek(0)
        response = HttpResponse(
            zip_buffer.getvalue(),
            content_type='application/zip'
        )
        
        zip_filename = f"Export_Performance_CU_{cu_type}_{date_debut.strftime('%Y%m%d')}_{date_fin.strftime('%Y%m%d')}.zip"
        response['Content-Disposition'] = f'attachment; filename="{zip_filename}"'
        
        return response
        
    except Exception as e:
        print(f"Erreur lors de la création du ZIP: {str(e)}")
        # En cas d'erreur, retourner seulement l'Excel
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        filename = f"Performance_CU_{cu_type}_{date_debut.strftime('%Y%m%d')}_{date_fin.strftime('%Y%m%d')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Sauvegarder le classeur dans la réponse
        wb.save(response)
        
        return response

def parse_log_file(log_file):
    """Analyse un fichier log et retourne un DataFrame avec les données structurées"""
    print(f"Analyse du fichier: {log_file}")
    
    # Expression régulière pour parser chaque ligne
    # Modifiée pour gérer les événements sans détails (comme MachineStop et MachineStart)
    log_pattern = re.compile(r"(\d{8} \d{2}:\d{2}:\d{2})\|@(\w+)(?:\s*:?\s*(.*))?")
    
    # Stockage des données
    data = []
    
    try:
        with open(log_file, "r", encoding="ISO-8859-1") as file:
            for line in file:
                match = log_pattern.match(line.strip())
                if match:
                    timestamp_str, event, details = match.groups()
                    
                    # Convertir le timestamp en objet datetime pour faciliter les calculs
                    timestamp = datetime.strptime(timestamp_str, "%Y%m%d %H:%M:%S")
                    data.append({
                        "Timestamp": timestamp,
                        "TimestampStr": timestamp_str,
                        "Event": event,
                        "Details": details.strip() if details else None
                    })
    except Exception as e:
        print(f"Erreur lors de la lecture du fichier {log_file}: {str(e)}")
        return None
        
    # Création du DataFrame
    if not data:
        print(f"Aucune donnée trouvée dans le fichier {log_file}")
        return None
        
    df = pd.DataFrame(data)
    return df

def analyze_machine_performance(df, log_file_name):
    """Analyse les performances d'une machine à partir des données de log"""
    if df is None or df.empty:
        return None
    
    # Extraire la date du log (pour l'affichage)
    log_date = df["Timestamp"].iloc[0].strftime("%Y-%m-%d")
    
    # Identifier le centre d'usinage (peut être extrait du nom de fichier ou des logs)
    # Pour l'instant, nous utilisons le nom du fichier
    cu_id = Path(log_file_name).stem
    
    # -----------------------
    # 1. Heure de première et dernière pièce
    # -----------------------
    stuk_df = df[df["Event"] == "StukUitgevoerd"]
    first_piece_time = stuk_df["Timestamp"].iloc[0] if not stuk_df.empty else None
    last_piece_time = stuk_df["Timestamp"].iloc[-1] if not stuk_df.empty else None
    
    # Total des pièces produites
    total_pieces = len(stuk_df)
    
    # Durée totale de production
    production_duration = None
    if first_piece_time and last_piece_time:
        production_duration = (last_piece_time - first_piece_time).total_seconds() / 3600  # en heures
    
    # -----------------------
    # 2. Temps d'attente
    # -----------------------
    wait_times = df[df["Event"] == "MachineWait"]["Details"].dropna()
    wait_timestamps = df[df["Event"] == "MachineWait"]["Timestamp"].dropna()
    
    total_wait_time = 0
    wait_periods = []
    
    for i, wait in enumerate(wait_times):
        try:
            # Format observé dans les logs: "XX sec" où XX est un nombre entier
            # Extraire les valeurs numériques (temps d'attente en secondes)
            wait_matches = re.findall(r"(\d+) sec", str(wait))
            wait_duration = 0
            
            if wait_matches:
                wait_duration = float(wait_matches[0])
            else:
                # Essayer aussi le format avec décimales au cas où
                decimal_matches = re.findall(r"(\d+\.\d+)", str(wait))
                if decimal_matches:
                    wait_duration = float(decimal_matches[0])
            
            if wait_duration > 0 and i < len(wait_timestamps):
                # Ajouter la période d'attente
                wait_start = wait_timestamps.iloc[i]
                wait_end = wait_start + pd.Timedelta(seconds=wait_duration)
                wait_periods.append({
                    "Start": wait_start,
                    "End": wait_end,
                    "Duration": wait_duration
                })
                total_wait_time += wait_duration
        except Exception as e:
            # Ignorer les erreurs silencieusement
            continue
    
    # Convertir en heures pour faciliter la compréhension
    total_wait_hours = total_wait_time / 3600
    
    # -----------------------
    # 3. Arrêts volontaires
    # -----------------------
    # Identifions les périodes d'arrêt (entre MachineStop et MachineStart)
    stop_events = df[df["Event"].isin(["MachineStop", "MachineStart"])].copy()
    
    # Extraire l'heure du dernier MachineStop
    machine_stop_events = df[df["Event"] == "MachineStop"]
    last_machine_stop = machine_stop_events["Timestamp"].iloc[-1] if not machine_stop_events.empty else None
    
    # Extraire l'heure du premier MachineStart
    machine_start_events = df[df["Event"] == "MachineStart"]
    first_machine_start = machine_start_events["Timestamp"].iloc[0] if not machine_start_events.empty else None
    
    total_stop_time = 0
    stop_periods = []
    
    if not stop_events.empty:
        stop_events = stop_events.reset_index()
        
        for i in range(len(stop_events) - 1):
            if stop_events.loc[i, "Event"] == "MachineStop" and stop_events.loc[i + 1, "Event"] == "MachineStart":
                stop_start = stop_events.loc[i, "Timestamp"]
                stop_end = stop_events.loc[i + 1, "Timestamp"]
                stop_duration = (stop_end - stop_start).total_seconds()
                stop_periods.append({
                    "Start": stop_start,
                    "End": stop_end,
                    "Duration": stop_duration
                })
                total_stop_time += stop_duration
    
    # Convertir en heures
    total_stop_hours = total_stop_time / 3600
    
    # Collecter les pièces produites dans une liste temporelle
    piece_events = []
    for idx, row in stuk_df.iterrows():
        piece_events.append({
            "Timestamp": row["Timestamp"],
            "Piece": row["Details"]
        })
    
    # -----------------------
    # 4. Références des barres débitées
    # -----------------------
    job_profiles = df[df["Event"] == "JobProfiel"]["Details"].dropna().tolist()
    
    # Extraire les références, longueurs et couleurs
    job_details = []
    for job in job_profiles:
        try:
            # Format attendu: R:215023 L:3359.0 C:7016 G ALU
            ref_match = re.search(r"R:(\w+)", job)
            length_match = re.search(r"L:(\d+\.\d+)", job)
            color_match = re.search(r"C:(\w+)", job)
            
            if ref_match and length_match:
                ref = ref_match.group(1)
                length = float(length_match.group(1))
                color = color_match.group(1) if color_match else "N/A"
                
                job_details.append({
                    "Reference": ref,
                    "Length": length,
                    "Color": color
                })
        except:
            continue
    
    # -----------------------
    # 5. Calcul des indicateurs de performance
    # -----------------------
    # Temps total disponible = durée entre première et dernière pièce
    if production_duration:
        # Temps de production effectif (heures)
        effective_production_time = production_duration - total_wait_hours - total_stop_hours
        
        # Temps total utilisé comme base de calcul (au lieu des 24h fixes)
        total_available_time = production_duration
        
        # Taux d'occupation de la machine
        occupation_rate = (effective_production_time / total_available_time) * 100
        
        # Taux d'attente
        wait_rate = (total_wait_hours / total_available_time) * 100
        
        # Taux d'arrêt volontaire
        stop_rate = (total_stop_hours / total_available_time) * 100
    else:
        effective_production_time = 0
        occupation_rate = 0
        wait_rate = 0
        stop_rate = 0
    
    # Construction du dictionnaire de résultats
    results = {
        "CU_ID": cu_id,
        "Date": log_date,
        "PremierePiece": first_piece_time,
        "DernierePiece": last_piece_time,
        "PremierMachineStart": first_machine_start,
        "DernierMachineStop": last_machine_stop,
        "TotalPieces": total_pieces,
        "DureeProduction": production_duration,
        "DureeProduction_hhmm": format_hours_minutes_hhmm(production_duration),
        "TempsAttente": total_wait_hours,
        "TempsAttente_hhmm": format_hours_minutes_hhmm(total_wait_hours),
        "TempsArretVolontaire": total_stop_hours,
        "TempsArretVolontaire_hhmm": format_hours_minutes_hhmm(total_stop_hours),
        "TempsProductionEffectif": effective_production_time,
        "TempsProductionEffectif_hhmm": format_hours_minutes_hhmm(effective_production_time),
        "TauxOccupation": occupation_rate,
        "TauxAttente": wait_rate,
        "TauxArretVolontaire": stop_rate,
        "JobDetails": job_details,
        "WaitPeriods": wait_periods,
        "StopPeriods": stop_periods,
        "PieceEvents": piece_events
    }
    
    return results

def format_hours_minutes(hours):
    """Convertit un nombre d'heures décimal en format 'Xh YYmin'"""
    if hours is None:
        return "N/A"
    
    h = int(hours)  # Partie entière (heures)
    m = int((hours - h) * 60)  # Partie décimale convertie en minutes
    
    return f"{h}h {m:02d}min"

def format_hours_minutes_hhmm(hours):
    """Convertit un nombre d'heures décimal en format 'hh:mm'"""
    if hours is None:
        return "N/A"
    
    h = int(hours)  # Partie entière (heures)
    m = int((hours - h) * 60)  # Partie décimale convertie en minutes
    
    return f"{h:02d}:{m:02d}"

def generate_fallback_charts(results):
    """Génère des graphiques avec matplotlib pour les performances d'une machine"""
    import matplotlib
    matplotlib.use('Agg')  # Utiliser le backend non-interactif
    import matplotlib.pyplot as plt
    import numpy as np
    import os
    from datetime import datetime
    import base64
    from io import BytesIO
    
    print("Génération des graphiques matplotlib")
    
    if not results:
        print("ERREUR: Aucun résultat fourni pour générer les graphiques")
        return None
    
    fallback_charts = {}
    
    try:
        # Définir un style plus moderne
        plt.style.use('ggplot')
        
        # GRAPHIQUE 1: CAMEMBERT - Répartition des Temps
        print("Génération du camembert")
        plt.figure(figsize=(8, 6))
        
        # Données pour le camembert
        labels = ['Production Effective', 'Attente', 'Arrêt Volontaire']
        sizes = [
            results.get('TempsProductionEffectif', 0),
            results.get('TempsAttente', 0),
            results.get('TempsArretVolontaire', 0)
        ]
        
        # Vérifier que les données ne sont pas toutes nulles
        if sum(sizes) == 0:
            sizes = [1, 0, 0]  # Valeurs par défaut pour éviter les erreurs
            print("ATTENTION: Toutes les valeurs sont nulles pour le camembert, utilisation de valeurs par défaut")
        
        colors = ['#4CAF50', '#FFC107', '#F44336']
        
        # Calculer les pourcentages
        total = sum(sizes)
        percentages = [round(size / total * 100, 1) if total > 0 else 0 for size in sizes]
        
        # Créer les labels avec pourcentages
        labels_with_pct = [f'{label}\n({pct}%)' for label, pct in zip(labels, percentages)]
        
        # Créer le camembert
        plt.pie(sizes, labels=labels_with_pct, colors=colors, autopct='', startangle=90, 
                wedgeprops={'edgecolor': 'white', 'linewidth': 1.5})
        plt.axis('equal')  # Equal aspect ratio ensures that pie is drawn as a circle
        plt.title('Répartition des Temps sur la Période d\'Utilisation', fontsize=14, pad=20)
        
        # Sauvegarder l'image en mémoire
        buffer = BytesIO()
        plt.savefig(buffer, format='png', dpi=100, bbox_inches='tight')
        buffer.seek(0)
        image_png = buffer.getvalue()
        buffer.close()
        
        # Encoder l'image en base64
        pie_chart_base64 = base64.b64encode(image_png).decode('utf-8')
        fallback_charts['pie_chart'] = f'data:image/png;base64,{pie_chart_base64}'
        plt.close()
        
        # GRAPHIQUE 2: BARRES - Indicateurs de Performance
        print("Génération du graphique en barres")
        plt.figure(figsize=(8, 6))
        
        # Données pour le graphique en barres - Utiliser les temps en minutes au lieu des pourcentages
        indicators = ['Temps Production', 'Temps Attente', 'Temps Arrêt']
        
        # Convertir les heures en minutes
        values_minutes = [
            results.get('TempsProductionEffectif', 0) * 60,  # Convertir en minutes
            results.get('TempsAttente', 0) * 60,             # Convertir en minutes
            results.get('TempsArretVolontaire', 0) * 60      # Convertir en minutes
        ]
        
        # Créer le graphique en barres
        bars = plt.bar(indicators, values_minutes, color=['#4CAF50', '#FFC107', '#F44336'], width=0.6)
        
        # Ajouter les valeurs au-dessus des barres (en format heures:minutes)
        for bar, value in zip(bars, values_minutes):
            # Convertir les minutes en format heures:minutes
            hours = int(value // 60)
            minutes = int(value % 60)
            time_str = f"{hours}h {minutes:02d}min"
            
            plt.text(bar.get_x() + bar.get_width()/2., value + 5,
                    time_str, ha='center', va='bottom', fontsize=11, fontweight='bold')
        
        # Ajuster l'échelle Y en fonction des valeurs
        max_value = max(values_minutes) if values_minutes else 0
        plt.ylim(0, max_value * 1.15)  # Ajouter 15% d'espace au-dessus pour les étiquettes
        
        plt.title('Indicateurs de Performance (Temps)', fontsize=14, pad=20)
        plt.ylabel('Temps (minutes)', fontsize=12)
        plt.grid(axis='y', linestyle='--', alpha=0.7)
        
        # Formater l'axe Y pour afficher les minutes de manière plus lisible
        from matplotlib.ticker import FuncFormatter
        
        def minutes_formatter(x, pos):
            hours = int(x // 60)
            minutes = int(x % 60)
            if hours > 0:
                return f"{hours}h{minutes:02d}"
            else:
                return f"{minutes}min"
        
        plt.gca().yaxis.set_major_formatter(FuncFormatter(minutes_formatter))
        
        # Sauvegarder l'image en mémoire
        buffer = BytesIO()
        plt.savefig(buffer, format='png', dpi=100, bbox_inches='tight')
        buffer.seek(0)
        image_png = buffer.getvalue()
        buffer.close()
        
        # Encoder l'image en base64
        bar_chart_base64 = base64.b64encode(image_png).decode('utf-8')
        fallback_charts['bar_chart'] = f'data:image/png;base64,{bar_chart_base64}'
        plt.close()
        
        # GRAPHIQUE 3: CHRONOLOGIE - Périodes d'Activité et d'Arrêt
        print("Génération de la chronologie")
        if results.get('PieceEvents') and len(results.get('PieceEvents', [])) > 0:
            fig, ax = plt.subplots(figsize=(12, 4))  # Réduire la hauteur de 6 à 4 pouces
            
            # Définir les couleurs
            colors = {
                'Production': '#4CAF50',
                'Attente': '#FFC107',
                'Arrêt': '#F44336'
            }
            
            # Créer trois lignes pour Production, Attentes et Arrêts
            y_positions = [3, 2, 1]  # Positions verticales pour chaque type d'événement
            labels = ['Production', 'Attentes', 'Arrêts']
            
            # Ajouter des rectangles pour les périodes de production
            if results.get('PremierePiece') and results.get('DernierePiece'):
                start_time = results['PremierePiece'].hour + results['PremierePiece'].minute/60
                end_time = results['DernierePiece'].hour + results['DernierePiece'].minute/60
                
                # Extraire les heures des événements
                piece_times = [event['Timestamp'].hour + event['Timestamp'].minute/60 for event in results.get('PieceEvents', [])]
                
                # Créer des blocs de production basés sur les pièces
                production_blocks = []
                current_block = {'start': start_time, 'end': start_time}
                
                # Trier les temps de pièces
                piece_times.sort()
                
                # Ajouter des points bleus pour chaque pièce produite
                for time in piece_times:
                    ax.scatter(time, y_positions[0], color='blue', s=30, zorder=5)
                
                # Ajouter des rectangles pour les périodes d'attente
                for period in results.get('WaitPeriods', []):
                    start = period['Start'].hour + period['Start'].minute/60
                    end = period['End'].hour + period['End'].minute/60
                    width = end - start
                    ax.add_patch(plt.Rectangle((start, y_positions[1]-0.4), width, 0.8, 
                                              color=colors['Attente'], alpha=0.7))
                
                # Ajouter des rectangles pour les périodes d'arrêt
                for period in results.get('StopPeriods', []):
                    start = period['Start'].hour + period['Start'].minute/60
                    end = period['End'].hour + period['End'].minute/60
                    width = end - start
                    ax.add_patch(plt.Rectangle((start, y_positions[2]-0.4), width, 0.8, 
                                              color=colors['Arrêt'], alpha=0.7))
                
                # Ajouter des rectangles verts pour les périodes de production
                # (entre les périodes d'attente et d'arrêt)
                production_periods = []
                current_time = start_time
                
                # Créer une liste combinée de toutes les périodes d'interruption
                all_interruptions = []
                for period in results.get('WaitPeriods', []):
                    all_interruptions.append({
                        'start': period['Start'].hour + period['Start'].minute/60,
                        'end': period['End'].hour + period['End'].minute/60,
                        'type': 'Attente'
                    })
                for period in results.get('StopPeriods', []):
                    all_interruptions.append({
                        'start': period['Start'].hour + period['Start'].minute/60,
                        'end': period['End'].hour + period['End'].minute/60,
                        'type': 'Arrêt'
                    })
                
                # Trier les interruptions par heure de début
                all_interruptions.sort(key=lambda x: x['start'])
                
                # Créer les périodes de production (entre les interruptions)
                if all_interruptions:
                    # Période de production avant la première interruption
                    if all_interruptions[0]['start'] > start_time:
                        ax.add_patch(plt.Rectangle((start_time, y_positions[0]-0.4), 
                                                  all_interruptions[0]['start'] - start_time, 0.8, 
                                                  color=colors['Production'], alpha=0.7))
                    
                    # Périodes de production entre les interruptions
                    for i in range(len(all_interruptions)-1):
                        if all_interruptions[i]['end'] < all_interruptions[i+1]['start']:
                            ax.add_patch(plt.Rectangle((all_interruptions[i]['end'], y_positions[0]-0.4), 
                                                      all_interruptions[i+1]['start'] - all_interruptions[i]['end'], 0.8, 
                                                      color=colors['Production'], alpha=0.7))
                    
                    # Période de production après la dernière interruption
                    if all_interruptions[-1]['end'] < end_time:
                        ax.add_patch(plt.Rectangle((all_interruptions[-1]['end'], y_positions[0]-0.4), 
                                                  end_time - all_interruptions[-1]['end'], 0.8, 
                                                  color=colors['Production'], alpha=0.7))
                else:
                    # S'il n'y a pas d'interruptions, toute la période est en production
                    ax.add_patch(plt.Rectangle((start_time, y_positions[0]-0.4), 
                                              end_time - start_time, 0.8, 
                                              color=colors['Production'], alpha=0.7))
            
            # Configurer les axes
            ax.set_yticks(y_positions)
            ax.set_yticklabels(labels)
            
            # Déterminer dynamiquement la plage horaire en fonction des données
            min_hour = 5  # Heure minimale par défaut
            max_hour = 22  # Heure maximale par défaut
            
            # Ajuster en fonction des données réelles si disponibles
            if results.get('PremierePiece') and results.get('DernierePiece'):
                # Calculer l'heure de début (arrondie à l'heure inférieure)
                start_hour = results['PremierePiece'].hour
                min_hour = max(5, start_hour - 1)  # Au moins 1h avant la première pièce, mais pas avant 5h
                
                # Calculer l'heure de fin (arrondie à l'heure supérieure)
                end_hour = results['DernierePiece'].hour
                if results['DernierePiece'].minute > 0:
                    end_hour += 1  # Ajouter une heure si des minutes sont présentes
                max_hour = min(23, max(18, end_hour + 0))  # Au moins jusqu'à 18h, au plus jusqu'à 22h
            
            ax.set_xlim(min_hour, max_hour)
            
            # Ajouter une grille horizontale
            ax.grid(axis='y', linestyle='--', alpha=0.3)
            
            # Créer le titre avec la date
            date_str = results.get('Date', '')
            if date_str:
                try:
                    # Convertir la date en format français
                    if isinstance(date_str, str):
                        date_obj = datetime.strptime(date_str, '%Y-%m-%d')
                    else:
                        date_obj = date_str
                    date_formatted = date_obj.strftime('%d/%m/%Y')
                    title_with_date = f'Chronologie des Périodes d\'Activité et d\'Arrêt - {date_formatted}'
                except:
                    title_with_date = f'Chronologie des Périodes d\'Activité et d\'Arrêt - {date_str}'
            else:
                title_with_date = 'Chronologie des Périodes d\'Activité et d\'Arrêt'
            
            # Ajouter des étiquettes et un titre
            plt.title(title_with_date, fontsize=14, pad=20)
            plt.xlabel('Heure', fontsize=12)
            
            # Ajouter une légende
            from matplotlib.patches import Patch
            legend_elements = [
                Patch(facecolor=colors['Production'], alpha=0.7, label='Production'),
                Patch(facecolor=colors['Attente'], alpha=0.7, label='Attente'),
                Patch(facecolor=colors['Arrêt'], alpha=0.7, label='Arrêt')
            ]
            # Placer la légende en haut à droite avec une taille de police réduite
            ax.legend(handles=legend_elements, loc='upper right', fontsize=9)
            
            # Réduire la taille du titre et des étiquettes pour s'adapter à la hauteur réduite
            plt.title(title_with_date, fontsize=12, pad=10)
            plt.xlabel('Heure', fontsize=10)
            
            # Réduire la taille des étiquettes des axes
            ax.tick_params(axis='both', which='major', labelsize=9)
            
            # Ajuster les marges pour optimiser l'espace
            plt.tight_layout()
            
            # Sauvegarder l'image en mémoire
            buffer = BytesIO()
            plt.savefig(buffer, format='png', dpi=100, bbox_inches='tight')
            buffer.seek(0)
            image_png = buffer.getvalue()
            buffer.close()
            
            # Encoder l'image en base64
            timeline_chart_base64 = base64.b64encode(image_png).decode('utf-8')
            fallback_charts['timeline_chart'] = f'data:image/png;base64,{timeline_chart_base64}'
            plt.close()
        else:
            print("Pas assez de données pour générer la chronologie")
            fallback_charts['timeline_chart'] = None
        
        print("Génération des graphiques matplotlib terminée avec succès")
        return fallback_charts
    except Exception as e:
        print(f"ERREUR lors de la génération des graphiques matplotlib: {str(e)}")
        import traceback
        traceback.print_exc()
        return None

def generate_comparison_chart(all_results):
    """Génère un graphique de comparaison Altair entre centres d'usinage"""
    import altair as alt
    import pandas as pd
    import random
    import string
    
    print("Début de la génération du graphique de comparaison")
    
    if not all_results:
        print("ERREUR: Aucun résultat fourni pour générer le graphique de comparaison")
        return None
    
    try:
        # Préparer les données
        comparison_data = []
        for result in all_results:
            comparison_data.append({
                'Centre': result['CU_ID'],
                'Taux Occupation': result['TauxOccupation'],
                'Date': result['Date']
            })
        
        print(f"DataFrame pour la comparaison créé: {len(comparison_data)} lignes")
        df_comparison = pd.DataFrame(comparison_data)
        
        # Créer le graphique de comparaison
        comparison_chart = alt.Chart(df_comparison).mark_bar().encode(
            x=alt.X('Centre:N', title='Centre d\'usinage'),
            y=alt.Y('Taux Occupation:Q', title='Taux d\'Occupation (%)', scale=alt.Scale(domain=[0, 100])),
            color=alt.Color('Centre:N', legend=None),
            tooltip=['Centre', 'Taux Occupation', 'Date']
        ).properties(
            width=800,
            height=350,
            title="Comparaison des Taux d'Occupation"
        )
        
        # Ajouter des étiquettes de valeur au-dessus des barres
        text = comparison_chart.mark_text(
            align='center',
            baseline='bottom',
            dy=-5
        ).encode(
            text=alt.Text('Taux Occupation:Q', format='.1f')
        )
        
        # Combiner les deux graphiques
        final_chart = (comparison_chart + text)
        
        # Générer le HTML avec les options d'intégration et modifier l'ID
        print("Génération du HTML pour le graphique de comparaison")
        comparison_html = final_chart.to_html(embed_options={'renderer': 'canvas', 'actions': False})
        print(f"HTML généré pour le graphique de comparaison: {len(comparison_html)} caractères")
        
        # Remplacer l'ID vis par comparison-vis dans le HTML généré
        comparison_html = comparison_html.replace('id="vis"', 'id="comparison-vis"')
        
        return comparison_html
    except Exception as e:
        print(f"ERREUR lors de la génération du graphique de comparaison: {str(e)}")
        import traceback
        traceback.print_exc()
        return None

def generate_comparison_chart_simple(all_results):
    """Version simplifiée du graphique de comparaison Altair entre centres d'usinage"""
    import altair as alt
    import pandas as pd
    
    print("Début de la génération simplifiée du graphique de comparaison")
    
    if not all_results:
        print("ERREUR: Aucun résultat fourni pour générer le graphique de comparaison")
        return None
    
    try:
        # Préparer les données
        comparison_data = []
        for result in all_results:
            comparison_data.append({
                'Centre': result['CU_ID'],
                'Taux Occupation': result['TauxOccupation']
            })
        
        df_comparison = pd.DataFrame(comparison_data)
        
        # Créer le graphique de comparaison
        comparison_chart = alt.Chart(df_comparison).mark_bar().encode(
            x='Centre:N',
            y='Taux Occupation:Q'
        ).properties(
            width=800,
            height=350,
            title="Comparaison des Taux d'Occupation"
        )
        
        # Générer le HTML
        comparison_html = comparison_chart.to_html()
        
        return comparison_html
    except Exception as e:
        print(f"ERREUR lors de la génération simplifiée du graphique de comparaison: {str(e)}")
        import traceback
        traceback.print_exc()
        return None


@login_required
def import_logs_view(request):
    """Vue pour importer les fichiers LOG via l'interface web"""
    from .services import LogImportService
    
    context = {
        'cu_types': [
            ('PVC', 'PVC (DEM12)'),
            ('ALU', 'ALU (DEMALU)'),
            ('HYBRIDE', 'HYBRIDE (SU12)')
        ]
    }
    
    if request.method == 'POST':
        cu_type = request.POST.get('cu_type')
        cu_name = request.POST.get('cu_name', '')
        import_type = request.POST.get('import_type')
        
        if not cu_type:
            messages.error(request, "Veuillez sélectionner un type de centre d'usinage.")
            return render(request, 'dashboard/import_logs.html', context)
        
        service = LogImportService()
        
        try:
            if import_type == 'auto_discover':
                # Import automatique
                base_log_directory = "logs"
                cu_directories = {
                    'PVC': 'DEM12 (PVC)',
                    'ALU': 'DEMALU (ALU)',
                    'HYBRIDE': 'SU12 (HYBRIDE)'
                }
                
                log_directory = os.path.join(base_log_directory, cu_directories[cu_type])
                
                if not os.path.exists(log_directory):
                    messages.warning(request, f'Répertoire non trouvé: {log_directory}')
                    return render(request, 'dashboard/import_logs.html', context)
                
                imported_sessions = service.import_directory(log_directory, cu_type)
                
                if imported_sessions:
                    messages.success(request, 
                        f'✓ {len(imported_sessions)} sessions importées avec succès pour {cu_type}'
                    )
                    context['imported_sessions'] = imported_sessions
                else:
                    messages.warning(request, f'Aucun fichier LOG trouvé dans {log_directory}')
                    
            elif import_type == 'file_upload':
                # Upload de fichier
                uploaded_file = request.FILES.get('log_file')
                
                if not uploaded_file:
                    messages.error(request, "Veuillez sélectionner un fichier LOG.")
                    return render(request, 'dashboard/import_logs.html', context)
                
                if not uploaded_file.name.endswith('.LOG'):
                    messages.error(request, "Le fichier doit avoir l'extension .LOG")
                    return render(request, 'dashboard/import_logs.html', context)
                
                # Sauvegarder temporairement le fichier
                import tempfile
                with tempfile.NamedTemporaryFile(delete=False, suffix='.LOG') as temp_file:
                    for chunk in uploaded_file.chunks():
                        temp_file.write(chunk)
                    temp_file_path = temp_file.name
                
                try:
                    session = service.import_log_to_database(
                        temp_file_path, 
                        cu_type, 
                        cu_name if cu_name else None
                    )
                    messages.success(request, f'✓ Session importée: {session}')
                    context['imported_sessions'] = [session]
                finally:
                    # Nettoyer le fichier temporaire
                    os.unlink(temp_file_path)
            
        except Exception as e:
            messages.error(request, f'Erreur lors de l\'import: {str(e)}')
    
    return render(request, 'dashboard/import_logs.html', context)